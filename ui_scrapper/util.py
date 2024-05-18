from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import os

def write_to_excel(data, filename):
    if not data:
        print("No data to write to Excel")
        return

    # write arrays to excel file
    # Initialize Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"

    # Header style
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header for Category, Kind, Product Name, Price, Image URL
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="Category")
    ws.cell(row=1, column=3, value="Brand")
    ws.cell(row=1, column=4, value="Product Name")
    ws.cell(row=1, column=5, value="Price")
    ws.cell(row=1, column=6, value="Image URL")
    ws.cell(row=1, column=7, value="Rating")
    ws.cell(row=1, column=8, value="Url")

    # Apply header style
    for col_num in range(1, 9):
        ws.cell(row=1, column=col_num).fill = header_fill
        ws.cell(row=1, column=col_num).alignment = header_alignment

    # Freeze the header
    ws.freeze_panes = ws['A2']  # Freeze the header row

    # Iterate through JSON data and write to Excel
    current_row = 2
    items_total = 0
    for product in data:
        ws.cell(row=current_row, column=1, value=product["id"])
        ws.cell(row=current_row, column=2, value=product["category"])
        ws.cell(row=current_row, column=3, value=product["brand"])
        ws.cell(row=current_row, column=4, value=product["name"])
        ws.cell(row=current_row, column=5, value=product["price"])
        ws.cell(row=current_row, column=6, value=product["image_url"])
        ws.cell(row=current_row, column=7, value=product["rating"])
        ws.cell(row=current_row, column=8, value=product["url"])
        current_row += 1
        items_total += 1

    # Adjust column widths to fit the longest data
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        if col_letter == 'C':  # For the Product Name column
            ws.column_dimensions[col_letter].width = 350 / 6  # 350 pixels
        else:
            ws.column_dimensions[col_letter].width = max_length + 2  # Add extra width for padding

    folder_path = os.path.dirname(filename)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print("Folder created:", folder_path)
    else:
        print("Folder already exists:", folder_path)

    wb.save(filename)
    

def display_progress(iteration, total, bar_length=50):
    percent = int(round((iteration / total) * 100))
    progress = int(round((bar_length * iteration) / total))
    bar = '#' * progress + '-' * (bar_length - progress)
    print(f'[{bar}] {percent}%\r', end='', flush=True)